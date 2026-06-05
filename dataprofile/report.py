# -*- coding: utf-8 -*-
"""
report.py - JSON and Excel report generation for DataProfiler.
"""

from __future__ import annotations

import io
import json
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

DARK_BLUE = "1F3864"
HEADER_FONT_COLOR = "FFFFFF"
RED_FILL = "FF4C4C"
YELLOW_FILL = "FFD966"
GREEN_FILL = "92D050"
LIGHT_BLUE_ALT = "DCE6F1"

_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=DARK_BLUE)


def _red_fill() -> PatternFill:
    return PatternFill("solid", fgColor=RED_FILL)


def _yellow_fill() -> PatternFill:
    return PatternFill("solid", fgColor=YELLOW_FILL)


def _green_fill() -> PatternFill:
    return PatternFill("solid", fgColor=GREEN_FILL)


def _apply_header_row(ws, row_num: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=10)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _pct_fill(pct: float, red_threshold: float, yellow_threshold: float) -> PatternFill | None:
    if pct > red_threshold:
        return _red_fill()
    if pct > yellow_threshold:
        return _yellow_fill()
    return _green_fill()


# ---------------------------------------------------------------------------
# JSON report
# ---------------------------------------------------------------------------

def build_json_report(results: list[dict[str, Any]]) -> str:
    """Return a formatted JSON string for the profiling run."""
    payload = {"profiling_run": results}
    return json.dumps(payload, indent=2, default=str)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def build_excel_report(results: list[dict[str, Any]]) -> bytes:
    """Return Excel bytes for all profiled tables."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _build_summary_sheet(wb, results)

    for result in results:
        _build_table_sheet(wb, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(wb: Workbook, results: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A2"

    headers = [
        "Table / File",
        "Total Records",
        "Total Columns",
        "Duplicate Count",
        "Duplicate %",
        "Quality Gate",
        "Errors",
        "Warnings",
    ]
    _apply_header_row(ws, 1, headers)

    for row_idx, r in enumerate(results, start=2):
        gate = r["quality_gate"]
        values = [
            r["table_name"],
            r["total_records"],
            r["total_columns"],
            r["duplicate_count"],
            r["duplicate_pct"],
            gate["status"],
            len(gate["errors"]),
            len(gate["warnings"]),
        ]
        alt_fill = PatternFill("solid", fgColor=LIGHT_BLUE_ALT) if row_idx % 2 == 0 else None

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 6:  # Quality gate column
                cell.fill = _red_fill() if val == "FAIL" else _green_fill()
                cell.font = Font(bold=True, size=10)
            elif alt_fill:
                cell.fill = alt_fill

    _auto_width(ws)
    ws.row_dimensions[1].height = 28


def _build_table_sheet(wb: Workbook, result: dict[str, Any]) -> None:
    raw_name = result["table_name"]
    sheet_name = raw_name[:31]  # Excel tab limit
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"

    # Meta block
    ws["A1"] = f"Table: {raw_name}"
    ws["A1"].font = Font(bold=True, size=12, color=DARK_BLUE)
    ws["A2"] = f"Total records: {result['total_records']}    |    " \
               f"Duplicates: {result['duplicate_count']} ({result['duplicate_pct']}%)    |    " \
               f"Quality Gate: {result['quality_gate']['status']}"
    ws["A2"].font = Font(bold=True, size=10)

    # Error / warning messages
    start_row = 3
    for err in result["quality_gate"]["errors"]:
        ws.cell(row=start_row, column=1, value=f"ERROR: {err}").font = Font(
            bold=True, color="CC0000", size=9
        )
        start_row += 1
    for warn in result["quality_gate"]["warnings"]:
        ws.cell(row=start_row, column=1, value=f"WARNING: {warn}").font = Font(
            bold=True, color="7F6000", size=9
        )
        start_row += 1

    header_row = start_row + 1

    # Column-level headers
    headers = [
        "Column Name",
        "Detected Type",
        "Is DOB",
        "Is Date Field",
        "Null Count",
        "Null %",
        "DOB > 100y Count",
        "DOB > 100y %",
        "DOB < 18y Count",
        "DOB < 18y %",
        "Date Stale (≥10y) Count",
        "Date Stale (≥10y) %",
    ]
    _apply_header_row(ws, header_row, headers)

    for row_idx, cp in enumerate(result["columns"], start=header_row + 1):
        alt_fill = PatternFill("solid", fgColor=LIGHT_BLUE_ALT) if row_idx % 2 == 0 else None

        row_vals = [
            cp["column"],
            cp["detected_type"],
            "Yes" if cp["is_dob"] else "No",
            "Yes" if cp["is_date"] else "No",
            cp["null_count"],
            cp["null_pct"],
            cp.get("dob_over_100_count", "N/A"),
            cp.get("dob_over_100_pct", "N/A"),
            cp.get("dob_under_18_count", "N/A"),
            cp.get("dob_under_18_pct", "N/A"),
            cp.get("date_stale_10y_count", "N/A"),
            cp.get("date_stale_10y_pct", "N/A"),
        ]

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Colour logic
            if col_idx == 6 and isinstance(val, (int, float)):  # null %
                cell.fill = _pct_fill(val, red_threshold=20, yellow_threshold=5)
            elif col_idx == 8 and isinstance(val, (int, float)):  # DOB >100y %
                cell.fill = _red_fill() if val > 0 else _green_fill()
            elif col_idx == 10 and isinstance(val, (int, float)):  # DOB <18y %
                cell.fill = _pct_fill(val, red_threshold=100, yellow_threshold=2)
            elif col_idx == 12 and isinstance(val, (int, float)):  # stale date %
                cell.fill = _pct_fill(val, red_threshold=80, yellow_threshold=40)
            elif alt_fill and col_idx not in (6, 8, 10, 12):
                cell.fill = alt_fill

    _auto_width(ws)
    ws.row_dimensions[header_row].height = 32
